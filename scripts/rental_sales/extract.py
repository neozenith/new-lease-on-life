#!/usr/bin/env python
# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "pandas",
#   "openpyxl",
#   "geopandas",
#   "pyarrow",
#   "ruamel.yaml",
#   "duckdb",
# ]
# ///
import argparse
import openpyxl as xl
from ruamel.yaml import YAML
from pathlib import Path
import pandas as pd
import logging
import geopandas as gpd
import datetime as dt
import duckdb

log = logging.getLogger(__name__)
SCRIPT_DIR = Path(__file__).parent.resolve()
CONFIG_FILE = SCRIPT_DIR / "extract_schema_mapping.yaml"
CONFIG = YAML().load(CONFIG_FILE.read_text())
PROJECT_DIR = SCRIPT_DIR.parent.parent.resolve()
# print(f"{json.dumps(CONFIG, indent=2)}")

# Geospatial references
SAL_FILE = PROJECT_DIR / "data/originals_converted/boundaries_victoria/SAL_2021_AUST_GDA2020_SHP/SAL_2021_AUST_GDA2020.parquet" # SAL_NAME21, SAL_CODE21
LGA_FILE = PROJECT_DIR / "data/originals_converted/boundaries_victoria/LGA_2024_AUST_GDA2020/LGA_2024_AUST_GDA2020.parquet" # LGA_NAME24, LGA_CODE24
POA_FILE = PROJECT_DIR / "data/originals_converted/boundaries_victoria/POA_2021_AUST_GDA2020_SHP/POA_2021_AUST_GDA2020.parquet" # POA_NAME21, POA_CODE21

SAL_GDF = gpd.read_parquet(SAL_FILE).to_crs(epsg=4326)
LGA_GDF = gpd.read_parquet(LGA_FILE).to_crs(epsg=4326)
POA_GDF = gpd.read_parquet(POA_FILE).to_crs(epsg=4326)

lgas = {r['LGA_NAME24'].lower().replace(' (vic.)', ''): r['LGA_CODE24'] for r in LGA_GDF[["LGA_NAME24", "LGA_CODE24"]].sort_values("LGA_NAME24").to_dict(orient="records")}
sal = {r['SAL_NAME21'].lower().replace(' (vic.)', ''): r['SAL_CODE21'] for r in SAL_GDF[["SAL_NAME21", "SAL_CODE21"]].sort_values("SAL_NAME21").to_dict(orient="records")}


lgas_used = set()
sal_used = set()

lga_not_found = set()
sal_not_found = set()

def normalise_name(filename:str, sheet_name: str) -> str:
    _filename = filename.lower().replace(" ", "_").replace(".xlsx", "")
    _sheet_name = sheet_name.lower().replace(" ", "_")
    return f"{_filename}__{_sheet_name}"

def cell_range_to_indices(cell_range: str):
    """Convert an Excel cell range (e.g., 'A1:C3') to zero-based row and column indices.

    Args:
        cell_range: Excel cell range in A1 notation
    """
    return xl.utils.cell.range_boundaries(cell_range)

def main(input_dir: Path, output_dir: Path, limit: int | None = None) -> None:
    """Extract rental and sales data from source files.

    Args:
        input_dir: Directory containing source files
        output_dir: Directory to save processed CSV files
        limit: Optional limit on number of files to process (for testing)
    """
    if output_dir is None:
        output_dir = input_dir.parent / "processed"
    output_dir.mkdir(parents=True, exist_ok=True)
    log.info(f"Extracting rental and sales data from {input_dir} to {output_dir}")

    configured_files = {Path(item["file"]).name: item for item in CONFIG}
    # Ensure output directory exists
    

    # Process each file in the input directory
    if input_dir.is_file():
        files = [input_dir]
    else:
        files = list(input_dir.rglob("*.xlsx"))
    
    records = []
    for file_path in files:
        log.info(f"Processing file: {file_path}")
        schema_map_for_file = configured_files.get(file_path.name)
        if not schema_map_for_file:
            log.info(f"  No schema mapping found for {file_path.name}, skipping.")
            continue

        results = process_file(file_path, schema_map_for_file, output_dir, limit)
        records.extend(results)

    df = pd.DataFrame(records)

    # Write CSV
    csv_path = output_dir / "all_extracted_data.csv"
    df.to_csv(csv_path, index=False)
    log.info(f"  - Wrote CSV: {csv_path}")

    # Write Parquet
    parquet_path = output_dir / "all_extracted_data.parquet"
    df.to_parquet(parquet_path, index=False)
    log.info(f"  - Wrote Parquet: {parquet_path}")

    # Write DuckDB
    db_path = output_dir / "all_extracted_data.duckdb"
    try:
        con = duckdb.connect(str(db_path))
        con.execute("DROP TABLE IF EXISTS rental_sales")
        con.execute("CREATE TABLE rental_sales AS SELECT * FROM df")
        row_count = con.execute("SELECT COUNT(*) FROM rental_sales").fetchone()[0]
        con.close()
        log.info(f"  - Wrote DuckDB: {db_path} ({row_count} rows)")
    except Exception as e:
        log.error(f"  - Failed to write DuckDB: {e}")

    log.info(f"Extracted a total of {len(df)} records across {len(files)} files.")
    log.info(f"Wrote files to {output_dir}")

def process_file(file_path: Path, schema_map_for_file: dict | None, output_dir: Path, limit: int | None = None):
    log.info("")
    log.info("----")
    log.info(f"Processing file: {file_path}")
    log.info("----")
    
    input_file_name = file_path.name
    log.info(f"Config for file: {schema_map_for_file}")
    workbook = xl.load_workbook(file_path, data_only=True)
    log.info(f"Workbook: {workbook}")
    log.info(f"Workbook: {list(workbook.sheetnames)}")

    configured_sheets = {item["sheet"]: item for item in schema_map_for_file["sheets"]}
    sheets = []
    for sheet in list(workbook.sheetnames):

        sheet_config = configured_sheets.get(sheet)
        if not sheet_config:
            log.info(f"  No config for sheet {sheet}, skipping.")
            continue

        sheet_obj = workbook[sheet]
        sheet_results = process_sheet(sheet_obj, schema_map_for_file, sheet_config, limit)
        
        output_file = output_dir / (normalise_name(input_file_name, sheet_obj.title) + ".csv")
        log.info(f"{input_file_name} - {sheet} normalised_name: {output_file}")
        sheet_df = pd.DataFrame(sheet_results)
        if not sheet_df.empty:
            
            log.info(f"  Writing {len(sheet_df)} rows to {output_file}")
            sheet_df.to_csv(output_file, index=False)
        sheets.extend(sheet_results)

    return sheets

def process_sheet(sheet_obj: xl.worksheet.worksheet.Worksheet, schema_map_for_file: dict, sheet_config: dict, limit: int | None = None):
    
    log.info(f"  Sheet: {sheet_obj.title}")
    log.info(f"  Config: {sheet_config}")
    if "time_bucket_range" in sheet_config:
        log.info(f"    Time bucket range: {sheet_config['time_bucket_range']} -> {cell_range_to_indices(sheet_config['time_bucket_range'])}")
    if "geospatial_range" in sheet_config:
        log.info(f"    Geospatial range: {sheet_config['geospatial_range']} -> {cell_range_to_indices(sheet_config['geospatial_range'])}")

    # iterate across timebucket columns and geospatial rows and cross reference the corresponding value cells
    # extracting the timebucket_value, geospatial_value, statistic_type, statistic_value
    time_bucket_start_col, time_bucket_start_row, time_bucket_end_col, time_bucket_end_row = cell_range_to_indices(sheet_config["time_bucket_range"])
    geo_start_col, geo_start_row, geo_end_col, geo_end_row = cell_range_to_indices(sheet_config["geospatial_range"])
    log.info(f"    Time bucket columns: {time_bucket_start_col} to {time_bucket_end_col}")
    log.info(f"    Geospatial rows: {geo_start_row} to {geo_end_row}")

    statistics = sheet_config["statistic"]
    log.info(f"    Statistics: {statistics}")

    rows = []
    time_bucket_format = schema_map_for_file.get("time_bucket_format")
    geospatial_type = schema_map_for_file.get("data_granularity")

    geo_lookup = {}
    if geospatial_type == "suburb":
        geo_lookup = sal
    elif geospatial_type == "lga":
        geo_lookup = lgas

    for geo_row in range(geo_start_row, geo_end_row + 1):
        geo_value = sheet_obj.cell(row=geo_row, column=geo_start_col).value
        if not geo_value or geo_value in ['Group Total', 'Grand Total', 'Victoria', 'Metro', 'Non-Metro']:
            log.info(f"      Row {geo_row} geospatial value is empty, skipping row.")
            continue

        geo_values = geo_value.lower().split("-") # Some values have codes appended with a dash

        if geo_value.lower() == "merri-bek":
            geo_values = ["merri-bek"] # Special case for Merri-bek which is spelt as "Merri-bek" and uses the delimiter
        if geo_value.lower() == "mornington penin'a":
            geo_values = ["mornington peninsula"] # Special case for Mornington Peninsula which has a contraction in the source data
        if geo_value.lower() == "colac-otway":
            geo_values = ["colac otway"] # Special case for Colac Otway which uses a dash in the source data
        if geo_value.lower() == "east brunswick":
            geo_values = ["brunswick east"] 
        if geo_value.lower() == "west brunswick":
            geo_values = ["brunswick west"] 
        if geo_value.lower() == "east st kilda":
            geo_values = ["st kilda east"] 
        if geo_value.lower() == "west st kilda":
            geo_values = ["st kilda west"] 

        geo_codes = [geo_lookup[v.strip()] for v in geo_values if v.strip() in geo_lookup]
        found = [v for v in geo_values if v.strip() in geo_lookup]
        not_found = [v for v in geo_values if v.strip() not in geo_lookup]

        if geospatial_type == "suburb":
            sal_used.update(found)
            sal_not_found.update(not_found)
        elif geospatial_type == "lga":
            lgas_used.update(found)
            lga_not_found.update(not_found)


        for time_col in range(time_bucket_start_col, time_bucket_end_col + 1):
            time_value = sheet_obj.cell(row=time_bucket_start_row, column=time_col).value
            if not time_value:
                log.info(f"      Column {time_col} time bucket value is empty, skipping column.")
                continue

            stat_index = (time_col - time_bucket_start_col) % len(statistics)
            stat_type = statistics[stat_index]
            
            stat_value = sheet_obj.cell(row=geo_row, column=time_col).value
            if stat_value is None or stat_value == "-" or stat_value == "":
                log.info(f"      Cell at row {geo_row}, column {time_col} is empty or invalid, skipping.")
                continue

            # Convert bedrooms to string to handle "all" values
            bedrooms_str = str(sheet_config["bedrooms"])

            # Convert stat_value to float to handle mixed types
            try:
                stat_value_float = float(stat_value)
            except (ValueError, TypeError):
                log.info(f"      Cell at row {geo_row}, column {time_col} has non-numeric value '{stat_value}', skipping.")
                continue

            row = {
                "geospatial": geo_value,
                "geospatial_codes": "-".join(geo_codes), # If multiple geo names are grouped together, then provide their codes joined by a dash too
                "geospatial_type": geospatial_type,
                "time_bucket": dt.datetime.strptime(str(time_value), time_bucket_format).date(),
                "dwelling_type": sheet_config["dwelling_type"],
                "bedrooms": bedrooms_str,
                "dwelling_class": f"{sheet_config['dwelling_type']}-{bedrooms_str}",
                "statistic": stat_type,
                "value": stat_value_float,
                "data_type": schema_map_for_file["data_type"],
                "data_frequency": schema_map_for_file["data_frequency"],
                "source_file": schema_map_for_file["file"],
                "source_sheet": sheet_obj.title,
                "cell": f"{xl.utils.get_column_letter(time_col)}{geo_row}"
            }
            rows.append(row)
            log.info(f"      Extracted row: {row}")

    geolookup_unused = set(geo_lookup.keys()).difference(sal_used if geospatial_type == "suburb" else lgas_used)
    print(f"""
    Geospatial type: {geospatial_type}
      sal_used: {sorted(sal_used)}
      sal_not_found: {sorted(sal_not_found)}
      lgas_used: {sorted(lgas_used)}
      lga_not_found: {sorted(lga_not_found)}
        geolookup_unused: {sorted(geolookup_unused)}
    """)
    return rows

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Extract rental and sales data from source files.")
    parser.add_argument("-v", "--verbose", action="store_true", help="Enable debug logging")
    parser.add_argument("-q", "--quiet", action="store_true", help="Show only errors")
    parser.add_argument("-f", "--force", action="store_true", help="Force reprocessing, ignoring cache")
    parser.add_argument("-n", "--dry-run", action="store_true", help="Run without writing outputs")
    parser.add_argument(
        "--input", type=Path, default=SCRIPT_DIR.parent / "data/originals/", help="Input directory containing source files."
    )
    parser.add_argument(
        "--output", type=Path, default=None, help="Output directory for processed CSV files."
    )
    parser.add_argument(
        "--limit", type=int, default=None, help="Limit the number of files to process (for testing)."
    )
    args = parser.parse_args()

    # Configure logging
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.ERROR if args.quiet else logging.INFO,
        format="%(asctime)s|%(name)s|%(levelname)s|%(filename)s:%(lineno)d - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )

    main(args.input, args.output, args.limit)
    # log.info(sal)